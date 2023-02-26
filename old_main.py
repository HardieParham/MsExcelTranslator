### 1 - Imports
from googletrans import Translator
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
print("1. Imports Loaded")


class Hardie:
    file_name = "test"
    translator = Translator()
    wb = load_workbook(f"input/{file_name}.xlsx")
    ws = wb.worksheets[0]
    # add_name = Hardie.translate(file_name)
    # wb.save(f"{file_name}_{add_name}_eng.xlsx")


    ### 15 - set settings (langauge in, languiage out)
    dest= "en"
    src= "de"


    char_to_replace = {'_': ' ', '/': ' ', '・': ' ', 'ｰ': ' ', '･' : ' ' , 'ｰ' : ' ', '･': ' ', '･': ' ', '*': ' ', '[': ' ', ']': ' ', '?': ' '}


    def __init__():
        pass


    ### 2 - Save new file
    def saving(file_name):
        Hardie.wb.save(file_name)
        print("2. Saved a New File")

    ### 2 - Save new file
    def initial():
        file_name = Hardie.file_name
        #add_name = Hardie.translate(file_name)
        new_name = f"{file_name}_eng.xlsx"
        print("2. New Name Created")
        return new_name




    ### 3 - create translating function
    def translate(text):
        try:
            for key, value in Hardie.char_to_replace.items():
                new_text = text.replace(key, value)
            #new_text = text.replace(("_", " "),("/", " "))
            translation = Hardie.translator.translate(new_text, dest= Hardie.dest, src= Hardie.src)
            print("3. Translation")
            return translation.text
        except:
            pass



    ## 4 - keep dict of worksheet names
    titles={}
    print("4. Titles Dict created")

    def first_wb_loop():
        for i, sheet in enumerate(Hardie.wb.worksheets):
            print(f"{i}, {sheet.title}")
            old_title = sheet.title
            print(old_title)
            new_title = Hardie.translate(sheet.title)
            print(new_title)
            if new_title == None:
                new_title = old_title
            if "/" in new_title:
                resulting = re.sub("/", " ", new_title)
                Hardie.wb.worksheets[i].title = resulting
                Hardie.titles[old_title] = resulting

            else:
                Hardie.wb.worksheets[i].title = new_title
                Hardie.titles[old_title] = new_title
        print("5. and 6. Titles Dict updated")
        print(Hardie.titles)






### 7 - function to find last cell in worksheet
    def lastcell():
        #Finds Last Entry in the last Column
        cols = (tuple(Hardie.ws.columns))
        print(cols)
        try:
            lcols = cols[-1]
            last = f"{lcols[-1]}"


            #Seperates last cell name from the string into row and col
            try:
                split = last.split("'.")
            except:
                print("WorkSheet Name Error! Change worksheet name to avoid the following character combination '.")
            position = split[1]
            row=""
            col=""
            for item in position:
                if item in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"):
                    col += item
                elif item == ">":
                    pass
                else:
                    row += item

            #Converts Column string into integer
            input = col.lower()
            output = []
            for character in input:
                number = ord(character) - 96
                output.append(number)
            if len(output) > 1:
                col_value = ((len(output) - 1) * 26) + output[-1]
            else:
                col_value = output[0]

            #Returns integers for Col and Row
            print("7. Found Last Cell in Sheet")
            return col_value, row

        except:
            return 0, 0



### 8 - translation loop through worksheet
    def sheet_trans_loop(lcol, lrow):
        print("8. Loop through sheet started")
        #Loop through the rows
        for row in range(1,(lrow+1)):
            #Nested loop through the columns first
            for col in range(1,(lcol+1)):
                #Set Column text character
                char = get_column_letter(col)
                #Extracts the value of the cell
                text = str(Hardie.ws[char + str(row)].value)


                ### 11 - if starts with =, check to see if worksheet names are in it. If yes, use dict translation. if not then pass
                if text[0] == "=":
                    translated = False
                    print(f"Checking if external sheet reference")
                    for key, value in Hardie.titles.items():
                        if key in text:
                            translation = text.replace(key, "'" + value + "'")
                            print(f"{char}{row} - {translation}")
                            Hardie.ws[char + str(row)].value = translation
                            translated = True
                    if translated == False:
                        print("was not translated (11)")


                ### 10 - if text, then translate
                elif text != "None":
                    translation = Hardie.translate(text)
                    print(f"{char}{row} - {translation} (10)")
                    Hardie.ws[char + str(row)].value = translation

                ### 9 - in text = none then pass
                else:
                    print(f"{char}{row} - none (9)")



    def wb_loop():
        for i, sheet in enumerate(Hardie.wb.worksheets):
            print("12. Go to next worksheet")
            print(f"{i}, {sheet.title}")
            Hardie.ws = Hardie.wb.worksheets[i]
            col_value = int(Hardie.lastcell()[0])
            row_value = int(Hardie.lastcell()[1])
            Hardie.sheet_trans_loop(lcol=col_value, lrow=row_value)
            #i += 1
            #Hardie.sheet_trans_loop(lcol = col_value, lrow = row_value)
            #Hardie.wb_loop()





### 16 - Start Loop
if __name__ == "__main__":
    new_name = Hardie.initial()
    Hardie.saving(new_name)
    Hardie.first_wb_loop()
    Hardie.wb_loop()
    print("Final Save")
    Hardie.saving(new_name)
    print("finished")










# Process
# 1 - Imports
# 2 - Save new file
# 3 - create translating function
# 4 - keep dict of worksheet names
# 5 - translate worksheet names
# 6 - update dict with translations
# 7 - function to find last cell in worksheet
# 8 - translation loop through worksheet
# 9 - in text = none then pass
# 10 - if text, then translate
# 11 - if starts with =, check to see if worksheet names are in it. If yes, use dict translationj. if not then pass
# 12 - go to next worksheet
# 13 - repeat finding cell and translate
# 14 - save file again
# 15 - set settings (langauge in, languiage out)
# 16 - Start Loop