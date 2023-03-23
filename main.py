# Standard imports
import re
import os
import logging


# External Imports
import googletrans
import openpyxl
import openpyxl.utils
from openpyxl.utils import get_column_letter


# Local Imports
from data.settings import data, char_to_replace




"""
Base Class for handling Text phrases.
"""
class Text():
    def __init__(self, dest, src):
        self.trans = googletrans.Translator()
        self.dest = dest
        self.src = src


    def translate(self, text):
        try:
            # There a few characters allowed in excel that break Google Translate's API
            # So first, need to remove any instances of these, and replace with a space
            for key, value in char_to_replace.items():
                new_text = text.replace(key, value)

            # Translate new API safe string
            translation = self.trans.translate(text=new_text, dest=self.dest, src=self.src)
            return translation.text
        
        except:
            logging.warning(f'Translation failed for {text}')




"""
Base Class to work with Excel Documents
Methods to Load and save documents
Methods to loop thru entire word documents to update Text Class
"""
class Workbook():
    def __init__(self, file):
        self.name = file
        self.new_name = 'ENG_' + file
        self.data = data
        self.wb = self.load_wb()
        self.ws_titles = {}
        self.trans = Text(dest=self.data['destination language'], src=self.data['source language'])


    def load_wb(self):
        wb = openpyxl.load_workbook(f"input/{self.name}")
        self.log_change(content=f'Starting translations for {self.name}')
        return wb


    def save_wb(self):
        self.wb.save(f'output/{self.new_name}')


    def log_change(self, content):
        with open('data/log.txt', 'a') as f:
            f.write(f'\n {content}')


    """
    Method to find the last cell (Row and column) that a value occurs
    NOTE: This is just the last column and last row that conatin a value. A value may not necessarily be at the row-column combination specified.
    e.g. Last item by column is in Z1, last item by row is in B 15, This function would return Z-15, even though there is not a value in that cell.
    """
    def get_lastcell(self):
            cols = (tuple(self.ws.columns))
            row=""
            col=""

            try:
                last_cell = str(cols[-1][-1])

                # Seperates the last cell's Letter Column and Number Row
                # NOTE Cell starts with a '.', so avoid '.' in the WS title 
                try:
                    split = last_cell.split("'.")
                except:
                    print("WorkSheet Name Error! Change worksheet name to avoid the following character combination '.")
                
                position = split[1]
                for item in position:
                    if item in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"):
                        col += item
                    elif item == ">":
                        pass
                    else:
                        row += item

                #Converting Column string into integer
                input = col.lower()
                output = []
                for character in input:
                    number = ord(character) - 96
                    output.append(number)

                # If Column has more than 1 letter, convert to correct integer
                if len(output) > 1:
                    col_value = ((len(output) - 1) * 26) + output[-1]
                else:
                    col_value = output[0]

                return col_value, row
            # If no cells were found, sheet is empty. Return zero coordinates.
            except:
                return 0, 0


    # Method to translate worksheet titles (the tabs at the bottom of excel)
    def translate_ws_titles(self):
        for i, sheet in enumerate(self.wb.worksheets):
            old_title = sheet.title
            new_title = self.trans.translate(sheet.title)
            
            # If translation failed, reset title to old_title
            if new_title == None:
                new_title = old_title

            # The character '/' in a WS title crashes excel, so need to remove it    
            if "/" in new_title:
                resulting = re.sub("/", " ", new_title)
                self.wb.worksheets[i].title = resulting
                self.ws_titles[old_title] = resulting
                self.log_change(content=f'Changed WS Title {old_title} to {new_title}')

            else:
                self.wb.worksheets[i].title = new_title
                self.ws_titles[old_title] = new_title
                self.log_change(content=f'Changed WS Title {old_title} to {new_title}')


    def loop_thru_worksheet(self, lcol, lrow):
        for row in range(1,(lrow+1)):
            for col in range(1,(lcol+1)):
                char = get_column_letter(col)
                text = str(self.ws[char + str(row)].value)

                # If content starts with '=', check to see if worksheet names are in the excel equation. If yes, use dict translation. if not then pass.
                if text[0] == "=":
                    translated_check = False
                    self.log_change(content=f'Checking if external sheet reference at {char},{row}')
                    for key, value in self.ws_titles.items():
                        if key in text:
                            translation = text.replace(key, ("'" + value + "'"))
                            self.log_change(content=f'{char}{row} was changed to {translation}')
                            self.ws[char + str(row)].value = translation
                            translated_check = True
 
                    if translated_check == False:
                        self.log_change(content=f'{char}{row}was not translated')

                elif text != "None":
                    translation = self.trans.translate(text)
                    self.log_change(content=f'{char}{row} - {translation}')
                    self.ws[char + str(row)].value = translation

                else:
                    self.log_change(content=f'{char}{row} - None')


    def loop_thru_document(self):
        self.translate_ws_titles()
        for i, sheet in enumerate(self.wb.worksheets):
            self.log_change(content=f'Starting to translate: {sheet}')
            print(f'Starting to translate: {sheet}')
            self.ws = self.wb.worksheets[i]
            (col_value, row_value) = self.get_lastcell()
            self.loop_thru_worksheet(lcol=int(col_value), lrow=int(row_value)) 

        self.log_change(content=f'Completed translating {sheet}')      
        print(f'Completed translating {self.name}')




"""
General Class for looping thru a folder of Excel documents
"""
class App():
    def __init__(self):
        self.input_path = './input'
        self.filename_list = self.get_files()

    def get_files(self):
        filelist = []
        dirpath = os.listdir(self.input_path)
        for file in dirpath:

            # Checks if file is valid (aka, not a subfolder)
            if os.path.isfile(os.path.join(self.input_path, file)):
                filelist.append(file)

        print(filelist)
        return filelist


    def delete_old_log(self):
        if os.path.exists('data/log.txt'):
            os.remove('data/log.txt')
        else:
            print('Log file not found')


    def main_loop(self):
        self.delete_old_log()
        for file in self.filename_list:
            print(f'\n Loading {file} \n')
            excel = Workbook(file)
            excel.loop_thru_document()
            excel.save_wb()




# Start main loop when file is run
if __name__ == "__main__":
    app = App()
    app.main_loop()